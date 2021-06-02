# -*- conding: utf-8 -*-
import os
import sys
import tkinter.ttk as ttk
import tkinter.messagebox as msgbox
import re
from tkinter import *
from tkinter import filedialog
from tkinter import scrolledtext
from openpyxl import load_workbook

os.chdir(os.path.dirname(os.path.abspath(__file__)))

root = Tk()
root.title("패스워드 변경 프로그램 version 0.5")
root.geometry("640x640")
root.minsize(width=640, height=640)
root.resizable(True, False)

head_lists = []
tail_lists = []

#================================ Function =====================================
def open_original_pw_file():
    original_pw_file_full_path = filedialog.askopenfilename(\
            initialdir=os.path.dirname(os.path.abspath(__file__)),\
            title="기존 패스워드 파일을 선택하세요",\
            filetypes=(("xlsx files", "*.xlsx"),("all files", "*.*"))\
                )
    global wb, ws, ws_copied

    if original_pw_file_full_path == "":
        pass
    else:
        lbl_original_pw_file_path.config(text=os.path.split(original_pw_file_full_path)[1])
        wb = load_workbook(original_pw_file_full_path)
        cmb_original_pw_sheet_name['value'] = wb.sheetnames
        cmb_original_pw_sheet_name.current(0)

def open_passvault_pw_file():
    passvault_pw_file_full_path = filedialog.askopenfilename(\
            initialdir=os.path.dirname(os.path.abspath(__file__)),\
            title="패스볼트 파일을 선택하세요",\
            filetypes=(("xlsx files", "*.xlsx"),("all files", "*.*"))\
                )
    global wb_PM
    if passvault_pw_file_full_path == "":
        pass
    else:
        lbl_passvault_pw_file_path.config(text=os.path.split(passvault_pw_file_full_path)[1])
        wb_PM = load_workbook(passvault_pw_file_full_path)
        cmb_passvault_pw_sheet_name['value'] = wb_PM.sheetnames
        cmb_passvault_pw_sheet_name.current(0)

# 패스워드 파일 변경
def run_change_pwasswd():
    if lbl_original_pw_file_path.cget("text") == "":
        msgbox.showwarning("오류", "패스워드 파일이 선택되지 않았습니다.")
        return

    head_lists.clear()
    for idx in range(1, 8):
        head_lists.append(root.children["passwd_frame"].children["first{}".format(idx)].get())

    tail_lists.clear()
    for idx in range(1, 8):
        tail_lists.append(root.children["passwd_frame"].children["end{}".format(idx)].get())

    errMsg = ""
    for idx, head_list in enumerate(head_lists):
        if head_list == '':
            errMsg += "\nType {}의 시작이 입력되지 않았습니다.".format(idx+1)

    for idx, tail_list in enumerate(tail_lists):
        if tail_list == '':
            errMsg += "\nType {}의 끝이 입력되지 않았습니다.".format(idx+1)

    if errMsg != "":
        msgbox.showwarning("오류", errMsg)
        return

    global wb, ws, ws_copied
    ws = wb[cmb_original_pw_sheet_name.get()]
    target = wb.copy_worksheet(ws)
    target.title = "ChangPW_Sheet"
    ws_copied = wb["ChangPW_Sheet"]

    progress = 0

    for x in range(4, ws.max_row + 1):
        pw_type = ws.cell(row=x, column=9).value
        if pw_type != "PM":
            change_passwd= get_passwd(ws.cell(row=x, column=8), ws.cell(row=x, column=9).value)
            ws_copied.cell(row=x, column=8).value = change_passwd
            if CheckVariety_1:
                ws_copied.cell(row=x, column=11).value = "echo \'{}:{}\' | chpasswd".format(ws.cell(row=x, column=7).value, change_passwd)
        progress = x / ws.max_row * 100
        p_var.set(progress)
        progress_bar.update()

    msgbox.showinfo("알림", "패스워드 변경이 완료되었습니다.\n저장버튼을 눌러야 파일로 저장됩니다.")

# 패스볼트 파일 변경
def run_passvault_pwasswd():
    if lbl_original_pw_file_path.cget("text") == "":
        msgbox.showwarning("오류", "패스워드 파일이 선택되지 않았습니다.")
        return

    if lbl_passvault_pw_file_path.cget("text") == "":
        msgbox.showwarning("오류", "PassVault 파일이 선택되지 않았습니다.")
        return
    try:
        global wb_PM, ws_PM, wb, ws, ws_copied
        ws_PM = wb_PM[cmb_passvault_pw_sheet_name.get()]

        progress = 0
        p_var.set(progress)
        progress_bar.update()

        for x in range(4, ws.max_row + 1):
            pw_type = ws.cell(row=x, column=9).value
            if pw_type == "PM":
                ws_copied.cell(row=x, column=8).value = get_PM_passwd(ws.cell(row=x, column=4))
            progress = x / ws.max_row * 100
            p_var.set(progress)
            progress_bar.update()

        msgbox.showinfo("알림", "PassVault 변경이 완료되었습니다.\n저장버튼을 눌러야 파일로 저장됩니다.")
    except EXCEPTION as err:
        msgbox.showwarning("오류", "패스워드 변경 부터 먼저 진행해주세요.")

#정해진 형식의 패스워드 생성을 위하여 아이피의 3,4번째 부분만 갖고 옴
def get_middle_value(c):
    result = get_merged_cell_value(c, ws)
    if result == "" or result == None:
        return result
    result = str(result).split("\n")[0] #첫 번째 줄만 갖고옴
    numbers = re.findall("\d+", result.strip())
    if len(numbers) == 4:
        result = numbers[2] + '.' + numbers[3]
    else:
        result = ""
    return result

# 패스워드 타입에 따라 패스워드를 생성
def get_passwd(c, type):
    global ws
    change_passwd = ""
    if type == "PM":
        change_passwd = ""
    elif type == "변경불가":
        change_passwd = c.value
    elif isinstance(type, int):
        mid_str = get_middle_value(ws.cell(row=c.row, column=4))
        if len(mid_str) <= 0:
            change_passwd = ""
        else:
            change_passwd = head_lists[int(type - 1)] + mid_str + tail_lists[int(type - 1)]
    else:
        change_passwd = c.value

    return change_passwd

# Merged Cell의 값 갖고 오기
def get_merged_cell_value(c, sheet):
    result = ""
    while True:
        if c.value:
            result = c.value
            break
        else:
            if c.row > 0:
                c = sheet.cell(row=c.row-1, column=c.column)
            else:
                result = ""
                break
    return result

# IP가 일치하는 지 확인
def ip_Match(ip1, ip2):
    result = False
    ws_ips = ip1.split("\n")
    ws_PM_ips = ip2.split("\n")

    for ws_ip in ws_ips:
        ws_ip = get_ip_from_ipstr(ws_ip)
        for ws_PM_ip in ws_PM_ips:
            ws_PM_ip = get_ip_from_ipstr(ws_PM_ip)
            if ws_ip == ws_PM_ip:
                return True
    return result

# IP 정보만을 추출
def get_ip_from_ipstr(ipstr):
    result = ""
    numbers = re.findall("\d+", ipstr.strip())
    if len(numbers) == 4:
        result = "{}.{}.{}.{}".format(numbers[0], numbers[1], numbers[2], numbers[3])

    return result

# 패스워드 파일의 셀정보로 패스볼트 파일의 패스워드 정보를 갖고옴
def get_PM_passwd(c):
    global ws_PM, ws
    result = ""
    ws_ip = get_merged_cell_value(c, ws)
    if ws_ip == "":
        return ""
    ws_id = ws.cell(row=c.row, column=7).value
    for x in range(3, ws_PM.max_row+1):
        pm_ip = get_merged_cell_value(ws_PM.cell(row=x, column=4), ws_PM)
        pm_id = ws_PM.cell(row=x, column=6).value
        if ws_id == pm_id and ip_Match(ws_ip, pm_ip):
             result = ws_PM.cell(row=x, column=7).value
    return result

# 파일저장
def run_save():
    save_text_file_fullpath = filedialog.asksaveasfilename(\
            initialdir=os.path.dirname(os.path.abspath(__file__)),\
            title="저장할 파일을 선택하세요",\
            filetypes=(("xlsx files", "*.xlsx"),("all files", "*.*"))\
                )
    global wb

    if save_text_file_fullpath != "":
        file_extension = ".xlsx"
        if os.path.splitext(save_text_file_fullpath)[1] != file_extension:
            save_text_file_fullpath = save_text_file_fullpath + file_extension
        wb.save(save_text_file_fullpath)
        wb.close()
        msgbox.showinfo("알림", "저장이 완료되었습니다.")
# 패스워드 타입 파일 열기
def open_pwtype():
    pwtype_full_path = filedialog.askopenfilename( \
        initialdir=os.path.dirname(os.path.abspath(__file__)), \
        title="패스워드 타입 파일을 선택해주세요", \
        filetypes=(("Text files", "*.txt"), ("all files", "*.*")) \
        )
    try:
        with open(pwtype_full_path, "r") as f:
            for line in f.readlines():
                contents = line.split("\t")
                if len(contents) < 3:
                    continue
                idx = contents[0][len(contents[0]) - 1:]
                root.children["passwd_frame"].children["first{}".format(idx)].delete(0, END)
                root.children["passwd_frame"].children["first{}".format(idx)].insert(0, contents[1].rstrip())
                root.children["passwd_frame"].children["end{}".format(idx)].delete(0, END)
                root.children["passwd_frame"].children["end{}".format(idx)].insert(0, contents[2].rstrip())
    except FileNotFoundError as err:
        pass

    pass
# 도움말 파일 열기
def show_help_msg():
    help_win = Toplevel(root)
    help_win.geometry("640x400")
    help_win.title("도움말")
    help_win.resizable(False, True)
    txtscroll = scrolledtext.ScrolledText(help_win, wrap='word')
    txtscroll.pack(fill="both", expand=True, padx=5, pady=5, ipady=4)
    try:
        with open("README.txt", "r", encoding='utf-8') as f:
            for line in f.readlines():
                txtscroll.insert(CURRENT, line)
    except FileNotFoundError as err:
        txtscroll.insert(END, "도움말 파일(README.txt)이 없습니다.")
    txtscroll.configure(state='disabled')
    close_btn = Button(help_win, text="닫기", width=10, command=help_win.withdraw)
    close_btn.pack(padx=5, pady=5, ipady=4)

#===================================== GUI =====================================
menu_bar = Menu(root)
root.config(menu = menu_bar)

file_menu = Menu(menu_bar, tearoff=0)
file_menu.add_command(label="패스워드타입 열기", command=open_pwtype)
file_menu.add_separator()
file_menu.add_command(label="종료", command=root.quit)

help_menu = Menu(menu_bar, tearoff=0)
help_menu.add_command(label="도움말", command=show_help_msg)
menu_bar.add_cascade(label="파일", menu=file_menu)
menu_bar.add_cascade(label="도움말", menu=help_menu)



original_pw_file_frame = LabelFrame(root, text="기존 패스워드 파일")
btn_open_original_pw_file = Button(original_pw_file_frame, width=10, text="열기", command=open_original_pw_file)
btn_open_original_pw_file.pack(side="left", padx=5, pady=5)

lbl_original_pw_file_path = Label(original_pw_file_frame, width=40)
lbl_original_pw_file_path.pack(side="left", fill="x", expand=True, padx=5, pady=5, ipady=4)

lbl_original_pw_sheet_name = Label(original_pw_file_frame, text="시트명", width=5)
lbl_original_pw_sheet_name.pack(side="left", padx=5, pady=5, ipady=4)

cmb_original_pw_sheet_name = ttk.Combobox(original_pw_file_frame, state="readonly", values=[], width=5)
cmb_original_pw_sheet_name.pack(side="right", fill="x", expand=True, padx=5, pady=5, ipady=4)

passvault_pw_file_frame = LabelFrame(root, text="패스볼트 파일")
btn_open_passvault_pw_file = Button(passvault_pw_file_frame, width=10, text="열기", command=open_passvault_pw_file)
btn_open_passvault_pw_file.pack(side="left", padx=5, pady=5)

lbl_passvault_pw_file_path = Label(passvault_pw_file_frame, width=40)
lbl_passvault_pw_file_path.pack(side="left", fill="x", expand=True, padx=5, pady=5, ipady=4)

lbl_passvault_pw_sheet_name = Label(passvault_pw_file_frame, text="시트명", width=5)
lbl_passvault_pw_sheet_name.pack(side="left", padx=5, pady=5, ipady=4)

cmb_passvault_pw_sheet_name = ttk.Combobox(passvault_pw_file_frame, state="readonly", values=[], width=5)
cmb_passvault_pw_sheet_name.pack(side="right", fill="x", expand=True, padx=5, pady=5, ipady=4)

passwd_rule_frame = LabelFrame(root, text="패스워드 규칙 설정", name="passwd_frame")
lbl_passwd_rule_first1 = Label(passwd_rule_frame, text="Type 1: 시작")
lbl_passwd_rule_first2 = Label(passwd_rule_frame, text="Type 2: 시작")
lbl_passwd_rule_first3 = Label(passwd_rule_frame, text="Type 3: 시작")
lbl_passwd_rule_first4 = Label(passwd_rule_frame, text="Type 4: 시작")
lbl_passwd_rule_first5 = Label(passwd_rule_frame, text="Type 5: 시작")
lbl_passwd_rule_first6 = Label(passwd_rule_frame, text="Type 6: 시작")
lbl_passwd_rule_first7 = Label(passwd_rule_frame, text="Type 7: 시작")

lbl_passwd_rule_end1 = Label(passwd_rule_frame, text="끝")
lbl_passwd_rule_end2 = Label(passwd_rule_frame, text="끝")
lbl_passwd_rule_end3 = Label(passwd_rule_frame, text="끝")
lbl_passwd_rule_end4 = Label(passwd_rule_frame, text="끝")
lbl_passwd_rule_end5 = Label(passwd_rule_frame, text="끝")
lbl_passwd_rule_end6 = Label(passwd_rule_frame, text="끝")
lbl_passwd_rule_end7 = Label(passwd_rule_frame, text="끝")

txt_passwd_rule_first1 = Entry(passwd_rule_frame, width=5, name="first1")
txt_passwd_rule_first2 = Entry(passwd_rule_frame, width=5, name="first2")
txt_passwd_rule_first3 = Entry(passwd_rule_frame, width=5, name="first3")
txt_passwd_rule_first4 = Entry(passwd_rule_frame, width=5, name="first4")
txt_passwd_rule_first5 = Entry(passwd_rule_frame, width=5, name="first5")
txt_passwd_rule_first6 = Entry(passwd_rule_frame, width=5, name="first6")
txt_passwd_rule_first7 = Entry(passwd_rule_frame, width=5, name="first7")

txt_passwd_rule_end1 = Entry(passwd_rule_frame, width=5, name="end1")
txt_passwd_rule_end2 = Entry(passwd_rule_frame, width=5, name="end2")
txt_passwd_rule_end3 = Entry(passwd_rule_frame, width=5, name="end3")
txt_passwd_rule_end4 = Entry(passwd_rule_frame, width=5, name="end4")
txt_passwd_rule_end5 = Entry(passwd_rule_frame, width=5, name="end5")
txt_passwd_rule_end6 = Entry(passwd_rule_frame, width=5, name="end6")
txt_passwd_rule_end7 = Entry(passwd_rule_frame, width=5, name="end7")

lbl_passwd_rule_explain1 = Label(passwd_rule_frame, text="관리자 계정")
lbl_passwd_rule_explain2 = Label(passwd_rule_frame, text="유지관리 계정")
lbl_passwd_rule_explain3 = Label(passwd_rule_frame, text="어플리케이션 계정")
lbl_passwd_rule_explain4 = Label(passwd_rule_frame, text="네트워크 en 계정")
lbl_passwd_rule_explain5 = Label(passwd_rule_frame, text="관리자 계정(외부)")
lbl_passwd_rule_explain6 = Label(passwd_rule_frame, text="유지관리 계정(외부)")
lbl_passwd_rule_explain7 = Label(passwd_rule_frame, text="어플리케이션 계정(외부)")

lbl_passwd_rule_first1.grid(column=1, row=0, padx=5, pady=5, ipady=4)
txt_passwd_rule_first1.grid(column=2, row=0, padx=5, pady=5, ipady=4)
lbl_passwd_rule_end1.grid(column=3, row=0, padx=5, pady=5, ipady=4)
txt_passwd_rule_end1.grid(column=4, row=0, padx=5, pady=5, ipady=4)
lbl_passwd_rule_explain1.grid(column=5, row=0, padx=5, pady=5, ipady=4, sticky="w")

lbl_passwd_rule_first2.grid(column=1, row=1, padx=5, pady=5, ipady=4)
txt_passwd_rule_first2.grid(column=2, row=1, padx=5, pady=5, ipady=4)
lbl_passwd_rule_end2.grid(column=3, row=1, padx=5, pady=5, ipady=4)
txt_passwd_rule_end2.grid(column=4, row=1, padx=5, pady=5, ipady=4)
lbl_passwd_rule_explain2.grid(column=5, row=1, padx=5, pady=5, ipady=4, sticky="w")

lbl_passwd_rule_first3.grid(column=1, row=2, padx=5, pady=5, ipady=4)
txt_passwd_rule_first3.grid(column=2, row=2, padx=5, pady=5, ipady=4)
lbl_passwd_rule_end3.grid(column=3, row=2, padx=5, pady=5, ipady=4)
txt_passwd_rule_end3.grid(column=4, row=2, padx=5, pady=5, ipady=4)
lbl_passwd_rule_explain3.grid(column=5, row=2, padx=5, pady=5, ipady=4, sticky="w")

lbl_passwd_rule_first4.grid(column=1, row=3, padx=5, pady=5, ipady=4)
txt_passwd_rule_first4.grid(column=2, row=3, padx=5, pady=5, ipady=4)
lbl_passwd_rule_end4.grid(column=3, row=3, padx=5, pady=5, ipady=4)
txt_passwd_rule_end4.grid(column=4, row=3, padx=5, pady=5, ipady=4)
lbl_passwd_rule_explain4.grid(column=5, row=3, padx=5, pady=5, ipady=4, sticky="w")

lbl_passwd_rule_first5.grid(column=1, row=4, padx=5, pady=5, ipady=4)
txt_passwd_rule_first5.grid(column=2, row=4, padx=5, pady=5, ipady=4)
lbl_passwd_rule_end5.grid(column=3, row=4, padx=5, pady=5, ipady=4)
txt_passwd_rule_end5.grid(column=4, row=4, padx=5, pady=5, ipady=4)
lbl_passwd_rule_explain5.grid(column=5, row=4, padx=5, pady=5, ipady=4, sticky="w")

lbl_passwd_rule_first6.grid(column=1, row=5, padx=5, pady=5, ipady=4)
txt_passwd_rule_first6.grid(column=2, row=5, padx=5, pady=5, ipady=4)
lbl_passwd_rule_end6.grid(column=3, row=5, padx=5, pady=5, ipady=4)
txt_passwd_rule_end6.grid(column=4, row=5, padx=5, pady=5, ipady=4)
lbl_passwd_rule_explain6.grid(column=5, row=5, padx=5, pady=5, ipady=4, sticky="w")

lbl_passwd_rule_first7.grid(column=1, row=6, padx=5, pady=5, ipady=4)
txt_passwd_rule_first7.grid(column=2, row=6, padx=5, pady=5, ipady=4)
lbl_passwd_rule_end7.grid(column=3, row=6, padx=5, pady=5, ipady=4)
txt_passwd_rule_end7.grid(column=4, row=6, padx=5, pady=5, ipady=4)
lbl_passwd_rule_explain7.grid(column=5, row=6, padx=5, pady=5, ipady=4, sticky="w")

CheckVariety_1= IntVar()
pw_script_ck1 = Checkbutton(passwd_rule_frame, text="패스워드 변경 스크립트 생성 여부", variable=CheckVariety_1)
pw_script_ck1.grid(column=6, row=0, padx=5, pady=5, ipady=4, sticky="e")

progress_frame = LabelFrame(root, text="진행상황")

p_var = DoubleVar()
progress_bar = ttk.Progressbar(progress_frame, maximum=100, variable=p_var)
progress_bar.pack(fill="x", padx=5, pady=5)

btn_frame = LabelFrame(root)
btn_run_change_pwasswd = Button(btn_frame, width=19, text="패스워드 변경", command=run_change_pwasswd)
btn_run_passvault_pwasswd = Button(btn_frame, width=19, text="PassVault 변경", command=run_passvault_pwasswd)
btn_run_save = Button(btn_frame, width=19, text="저장", command=run_save)
btn_run_close = Button(btn_frame, width=19, text="닫기", command=root.quit)

btn_run_change_pwasswd.pack(side=LEFT, fill=BOTH, expand=YES, padx=10, pady=10, ipady=4)
btn_run_passvault_pwasswd.pack(side=LEFT, fill=BOTH, expand=YES, padx=10, pady=10, ipady=4)
btn_run_save.pack(side=LEFT, fill=BOTH, expand=YES, padx=10, pady=10, ipady=4)
btn_run_close.pack(side=LEFT, fill=BOTH, expand=YES, padx=10, pady=10, ipady=4)

#===================== 위젯배치 ===========================
original_pw_file_frame.pack(side=TOP, fill=BOTH, expand=YES, padx=10, pady=10)
passvault_pw_file_frame.pack(side=TOP, fill=BOTH, expand=YES, padx=10, pady=10)
passwd_rule_frame.pack(side=TOP, fill=BOTH, expand=YES, padx=10, pady=10)
progress_frame.pack(side=TOP, fill=BOTH, expand=YES, padx=10, pady=10)
btn_frame.pack(side=TOP, fill=BOTH, expand=YES, padx=10, pady=10)

try:
    with open("pwtype.txt", "r") as f:
        for line in f.readlines():
            contents = line.split("\t")
            if len(contents) < 3:
                continue
            idx = contents[0][len(contents[0])-1:]
            root.children["passwd_frame"].children["first{}".format(idx)].insert(0, contents[1].rstrip())
            root.children["passwd_frame"].children["end{}".format(idx)].insert(0, contents[2].rstrip())
except FileNotFoundError as err:
    pass

root.mainloop()