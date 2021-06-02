from openpyxl import load_workbook
import re
wb = load_workbook("D:\\PW\\PW_sample.xlsx")
wb_PM = load_workbook("D:\\PW\\PassVault_sample.xlsx")
ws = wb["원본"]
ws_PM = wb_PM.active

target = wb.copy_worksheet(ws)
target.title = "ChangPW_Sheet"

ws_copied = wb["ChangPW_Sheet"]

head_list = ["Th","Rn","Tl","Eh","Dls", "Cjs","Tl"] #때꾸씨또, 인천씨
tail_list = ['&*', '()', '_+','&*','&*', '()', '_+']

#정해진 형식의 패스워드 생성을 위하여 아이피의 3,4번째 부분만 갖고 옴
def get_middle_value(c):
    result = get_merged_cell_value(c, ws)
    if result == "":
        return result
    result = result.split("\n")[0] #첫 번째 줄만 갖고옴
    numbers = re.findall("\d+", result.strip())
    if len(numbers) == 4:
        result = numbers[2] + '.' + numbers[3]
    else:
        result = ""
    return result

# 패스워드 타입에 따라 패스워드를 생성
def get_passwd(c, type):
    change_passwd = ""
    if type == "PM":
        change_passwd = ""
    elif type == "변경불가":
        change_passwd = c.value
    elif isinstance(type, int):
        mid_str = get_middle_value(ws.cell(row=c.row, column=4))
        if len(mid_str) <= 0:
            change_passwd = ""
        else:
            change_passwd = head_list[int(type - 1)] + mid_str + tail_list[int(type - 1)]
    else:
        change_passwd = c.value

    return change_passwd

# Merged Cell의 값 갖고 오기
def get_merged_cell_value(c, sheet):
    result = ""
    while True:
        if c.value:
            result = c.value
            break
        else:
            if c.row > 0:
                c = sheet.cell(row=c.row-1, column=c.column)
            else:
                result = ""
                break
    return result

# IP가 일치하는 지 확인
def ip_Match(ip1, ip2):
    result = False
    ws_ips = ip1.split("\n")
    ws_PM_ips = ip2.split("\n")

    for ws_ip in ws_ips:
        ws_ip = get_ip_from_ipstr(ws_ip)
        for ws_PM_ip in ws_PM_ips:
            ws_PM_ip = get_ip_from_ipstr(ws_PM_ip)
            if ws_ip == ws_PM_ip:
                return True
    return result

# IP 정보만을 추출
def get_ip_from_ipstr(ipstr):
    result = ""
    numbers = re.findall("\d+", ipstr.strip())
    if len(numbers) == 4:
        result = "{}.{}.{}.{}".format(numbers[0], numbers[1], numbers[2], numbers[3])

    return result


# 패스워드 파일의 셀정보로 패스볼트 파일의 패스워드 정보를 갖고옴
def get_PM_passwd(c):
    result = ""
    ws_ip = get_merged_cell_value(c, ws)
    if ws_ip == "":
        return ""
    ws_id = ws.cell(row=c.row, column=7).value
    for x in range(3, ws_PM.max_row+1):
        pm_ip = get_merged_cell_value(ws_PM.cell(row=x, column=4), ws_PM)
        pm_id = ws_PM.cell(row=x, column=6).value
        if ws_id == pm_id and ip_Match(ws_ip, pm_ip):
             result = ws_PM.cell(row=x, column=7).value
    return result

for x in range(4, ws.max_row + 1):
    pw_type = ws.cell(row=x, column=9).value
    if pw_type == "PM":     # 타입이 PM일 경우 PassVault의 암호를 갖고옴
        ws_copied.cell(row=x, column=8).value = get_PM_passwd(ws.cell(row=x, column=4))
    else:                   # 나머지 타입은 규칙에 맞게 변경
        ws_copied.cell(row=x, column=8).value = get_passwd(ws.cell(row=x, column=8),ws.cell(row=x, column=9).value)

wb.save("D:\\PW\\PW_sample_test.xlsx")
wb.close()




