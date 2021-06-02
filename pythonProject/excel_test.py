from openpyxl import Workbook
from random import *
# wb = Workbook() # 새 워크북 생성
# ws = wb.active # 현재 활성화된 sheet 가져옴
# ws.title = "title Test" # sheet의 이름을 변경
# wb.save("sample.xlsx")
# wb.close()
#=============================== sheet ============================
# wb = Workbook()
# ws = wb.create_sheet() # 새로운 시트를 생성
# ws.title = "My Sheet" # sheet의 이름을 변경
# ws.sheet_properties.tabColor = "ff66ff" #RGB 형태로 값을 변경 (구글에서 RGB 검색 w3schools.com)
#
# ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
# ws2 = wb.create_sheet("Sheet_3",0) # index의 위치에 sheet 생성
#
# new_ws = wb["Sheet_3"] # Dict 형태로 sheet 에 접근
#
# print(wb.sheetnames) # 모든 Sheet 이름 확인
#
# #Sheet 복사
# new_ws["A1"] = "Test"
# target = wb.copy_worksheet(new_ws)
# target.title = "Copied Sheet"
#
# wb.save("sample.xlsx")
# wb.close()
#=============================== cell ============================
# wb = Workbook()
# ws = wb.active # 현재 활성화된 sheet 가져옴
# ws.title = "Test"
# ws["A1"] = 1
# ws["A2"] = 2
# ws["A3"] = 3
#
# ws["B1"] = 4
# ws["B2"] = 5
# ws["B3"] = 6
#
# print(ws["A1"]) # A1 셀의 정보를 출력
# print(ws["A1"].value) # A1 셀의 값을 출력
# print(ws["A10"].value) # 값이 없을 땐 None 출력
# #row = 1, 2, 3
# #column = A, B, C
# print(ws.cell(row=1, column=2).value) # ws["B1"].value
#
# ws.cell(column=3, row=1, value=10)
#
# for x in range(1, 11): # 10개 row
#     for y in range(1, 11): # 10개 column
#         ws.cell(row=x, column=y, value=randint(0, 100)) #0~100 사이의 숫자
#
# wb.save("sample.xlsx")
# wb.close()
#=============================== open_file ============================
# from openpyxl import load_workbook # 파일 불러오기
# wb = load_workbook("sample.xlsx") #sample.xlsx 파일에서 wb을 불러옴
# ws =  wb.active # 활성화 된 Sheet
#
# # for x in range(1, 11):
# #     for y in range(1, 11):
# #         print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4 ..
# #     print()
#
# # cell 갯수를 모를 때
# for x in range(1, ws.max_row + 1):
#     for y in range(1, ws.max_column + 1):
#         print(ws.cell(row=x, column=y).value, end=" ")  # 1 2 3 4 ..
#     print()
# wb = Workbook()
# ws = wb.active
#
# #1줄씩 데이터 넣기
# ws.append(["번호", "영어", "수학"])
# for i in range(1, 11): #10개 데이터 넣기
#     ws.append([i,randint(0, 100), randint(0, 100)])

# col_B = ws["B"]
# for cell in col_B:
#     print(cell.value, end=" ")

#col_range = ws["B:C"] #영어, 수학 column 함께 가지고 오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value, end=" ")
#     print()
# row_title = ws[1] # 1번째 row만 가지고 오기
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6] # 2번째 줄에서 6번째 줄까지 가지고 오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# from openpyxl.utils.cell import coordinate_from_string
#
# row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         #print(cell.value, end= " ")
#         #print(cell.coordinate, end=" ")
#         xy = coordinate_from_string(cell.coordinate)
#         #print(xy, end= " ")
#         print(xy[0], end="") # A
#         print(xy[1], end=" ") # 1
#     print()
# 전체 rows
#print(tuple(ws.rows))
# 전체 columns
#print(tuple(ws.columns))

# for row in tuple(ws.rows):
#     print(row[2].value)

# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows(): #전체 row
#     print(row)
#     print(row[2].value)

# 1번째 줄부터 5번째 줄까지, 2번째 열부터 3번째 열까지
# for column in ws.iter_cols(min_row=2, max_row=11, min_col=2, max_col=3): #전체 column
#     print(column[0].value, column[1].value)

# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3): #범위 지정
#     print(row[0].value, row[1].value)
#=============================== search_file ============================
# from openpyxl import load_workbook # 파일 불러오기
# wb = load_workbook("sample.xlsx")
# ws = wb.active
#
# # for row in ws.iter_rows(min_row=2):
# #     #번호, 영어, 수학
# #     if int(row[1].value) > 80:
# #         print(row[0].value, "번 학생은 영어 천재: {}점".format(row[1].value))
#
# # for row in ws.iter_rows(max_row=1):
# #     for cell in row:
# #         if cell.value == "영어":
# #             cell.value = "컴퓨터"
#
# #wb.save("sample_modify.xlsx")

#=============================== insert_file ============================
# from openpyxl import load_workbook # 파일 불러오기
# wb = load_workbook("sample.xlsx")
# ws = wb.active

#ws.insert_rows(8) #8번째 줄이 비워짐
#ws.insert_rows(8, 5) #8번째 줄 위치에 5줄을 추가

#ws.insert_cols(2) #B번째 열이 비워짐(새로운 빈 열이 추가)
#ws.insert_cols(2, 3) #B번째 열에서 5열이 추가

#=============================== delete_file ============================
# from openpyxl import load_workbook # 파일 불러오기
# wb = load_workbook("sample.xlsx")
# ws = wb.active

#ws.delete_rows(6) #6번째 줄이 비워짐
#ws.delete_rows(6, 3) #6번째 줄 위치에서 3줄을 삭제
#ws.delete_cols(2) #2번째 열 삭제
#ws.delete_cols(2, 3) #2번째 열부터 3열 삭제
#=============================== move_file ============================
# from openpyxl import load_workbook # 파일 불러오기
# wb = load_workbook("sample.xlsx")
# ws = wb.active

# 번호 영어 수학
# 번호 (국어) 영어 수학

# ws.move_range("B1:C11", rows=0, cols=1)
# ws["B1"].value = "국어" # B1 셀에 '국어' 입력
# 번호 영어 수학 지정한 위치로 이동
#ws.move_range("C1:C11", rows=5, cols=-1)
#wb.save("sample_move.xlsx")

#wb.save("sample.xlsx")

#=============================== chart_file ============================
# from openpyxl import load_workbook # 파일 불러오기
# from openpyxl.chart import BarChart, Reference, LineChart
#
# wb = load_workbook("sample.xlsx")
# ws = wb.active
# # bar_value = Reference(ws, min_row=2, max_row=11, min_col=2, max_col=3)
# # bar_chart = BarChart()
# # bar_chart.add_data(bar_value)
#
# # 제목을 포함
# line_value = Reference(ws, min_row=1, max_row=11, min_col=2, max_col=3)
# line_chart = LineChart()
# line_chart.add_data(line_value, titles_from_data=True) # 계열 > 영어, 수학 (제목에서 갖고옴)
# line_chart.title = "성적표" # 제목
# line_chart.style = 10 #미리 정의된 스타일을 적용, 사용자가 개별 지정도 가능
# line_chart.y_axis.title = "점수"
# line_chart.x_axis.title = "번호"
# #구글에서 openpyxl 검색
#
# ws.add_chart(line_chart, "E1") #차트 넣을 위치 정의
# wb.save("sample_chart.xlsx")

#wb.close()

